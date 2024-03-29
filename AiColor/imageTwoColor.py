import cv2 as cv
import numpy as np
import matplotlib.pyplot as plt
import PIL
import pandas as pd
import os, glob
import os.path
import os
import sys
import glob
import openpyxl
# pip install sklearn # sklearn 설치 안되어 있을 시 이거 설치 후 실행(VS CODE에서는 command 창에서)
from sklearn.cluster import KMeans

# 워크북(엑셀파일)을 새로 만듭니다.
wb = openpyxl.Workbook()

# 현재 활성화된 시트를 선택합니다.
sheet = wb.active

# 두 번째 시트작성(시트이름 2nd sheet로 설정)
# sheet2 = wb.create_sheet("2nd sheet")

# sheet1의 시트이름 변경
sheet.title = "color_code"

r = g = b = 0 #0값으로 초기화
index=["color","color_name","hex","R","G","B"]
# 파일 불러오기
# colors.cvs - 137가지의 색상에 대한 각 R,G,B 값을 가지고 있는 엑셀
csv = pd.read_csv('colors.csv', names=index, header=None) 

#모든 색상으로부터 최소 거리를 계산하고 가장 일치하는 색상을 얻는 함수(가장 가까운 색 구하기)
def getColorName(R,G,B): 
    minimum = 10000
    for i in range(len(csv)):
        d = abs(R- int(csv.loc[i,"R"])) + abs(G- int(csv.loc[i,"G"]))+ abs(B- int(csv.loc[i,"B"]))
        if(d<=minimum):
            minimum = d
            cname = csv.loc[i,"color_name"]
    return cname


def readImage() :
    img_files = glob.glob('./mantoman/*.jpg')
    # 슬라이드 쇼 반복을 위한 반복문
    count = len(img_files)
    index = 0

    while True:
        img = cv.imread(img_files[index])
        img2 = img_files[index]

        # 예외처리
        if img is None:     
            print("이미지를 불러오는데 실패했습니다.")
            break
            
        # 경로출력
        print(img2)
        
        # cvtColor():사진의 색깔을 바꿔주는 함수
        img = cv.cvtColor(img, cv.COLOR_BGR2RGB)


        # 절대크기(너비,높이)
        dim = (500, 300)

        #이미지 크기 변환 함수: cv.resize(입력 이미지, 절대크기,보간법)
        img = cv.resize(img, dim, interpolation = cv.INTER_AREA) #cv.INTER_AREA: 영역보간법

        clt= KMeans(n_clusters = 2) # 2개 클러스터로 모델 생성

        # fit: 모델을 학습시켜 모델 객체에 저장하는 함수
        #reshape(행,열): 행,열로 된 2차원 배열을 재구성하는 함수 
        clt.fit(img.reshape(-1, 3)) 

        clt.labels_

        # 가장 높은 비율 2가지 색에 대한 각각의 RGB 값을 구함 
        clt.cluster_centers_

        # clt.cluster_centers_를 arr라는 변수에 대입
        arr = clt.cluster_centers_

        # 배열을 차례대로 R G B 값에 맞게 넣어줌
        # 그 이후에 round 함수를 통해 반올림 해주고 그 값을 정수로 변환
        r1 = arr[0][0]
        g1 = arr[0][1]
        b1 = arr[0][2]

        text1 = getColorName(r1,g1,b1)

        # 배열을 차례대로 R G B 값에 맞게 넣어줌
        # 그 이후에 round 함수를 통해 반올림 해주고 그 값을 정수로 변환
        r2 = arr[1][0]
        g2 = arr[1][1]
        b2 = arr[1][2]

        text2 = getColorName(r2,g2,b2)

        sheet.cell(row=index+1, column=1).value = text1  #sheet에 엑셀 1열에 text1값 대입
        sheet.cell(row=index+1, column=2).value = text2  #sheet에 엑셀 2열에 text1값 대입

        # (엑셀파일)을 원하는 이름으로 저장. -> "imageTwoColor.xlsx"로 지정
        # 단 2가지 색으로 압축한 것이기 때문에 추후에 1가지 색으로 선정하여 저장할 때는
        # 확장자를 csv로 해주어야 testColor.py에서 문제 없이 수행될 수 없음
       
        wb.save('imageTwoColor.xlsx')

        # index가 이미지 리스트보다 커지거나 같아지면 break
        index += 1      
        if count == index :
            break

readImage()